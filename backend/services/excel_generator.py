import io
import re
import zipfile
import base64
from pathlib import Path
from datetime import datetime
from typing import List
from openpyxl import load_workbook
from copy import copy


def set_cell_value(ws, cell_ref: str, value):
    """
    Safely set a cell value, handling merged cells.
    For merged cells, writes to the top-left cell of the merge range.
    """
    cell = ws[cell_ref]

    # Check if this cell is part of a merged range
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Get the top-left cell of the merged range
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            top_left.value = value
            return

    # Not merged, set directly
    cell.value = value


# Excel template row mappings and limits
TRAVEL_ROWS = list(range(13, 19))  # Rows 13-18 for travel general
MILEAGE_ROWS = list(range(23, 27))  # Rows 23-26 for car mileage
HOSPITALITY_ROWS = list(range(32, 36))  # Rows 32-35 for hospitality
OTHER_ROWS = list(range(41, 48))  # Rows 41-47 for other expenses

MAX_TRAVEL = len(TRAVEL_ROWS)  # 6
MAX_MILEAGE = len(MILEAGE_ROWS)  # 4
MAX_HOSPITALITY = len(HOSPITALITY_ROWS)  # 4
MAX_OTHER = len(OTHER_ROWS)  # 7

# Map active_section values to Excel sections
SECTION_TO_EXCEL = {
    "travel_general": "travel",
    "travel_mileage": "mileage",
    "hospitality": "hospitality",
    "other": "other",
}


def generate_expense_filename(parsed: dict, original_filename: str) -> str:
    """Generate renamed filename based on parsed data (new field structure)."""
    active_section = parsed.get("active_section", "other")
    fields = parsed.get("fields", {})
    section_data = fields.get(active_section, {})

    # Determine expense type based on section
    if active_section == "travel_general":
        expense_type = section_data.get("mode", "TRAVEL")
    elif active_section == "travel_mileage":
        expense_type = "MILEAGE"
    elif active_section == "hospitality":
        expense_type = "HOSPITALITY"
    else:
        expense_type = section_data.get("expense_type", "OTHER")

    expense_type = (expense_type or "OTHER").upper().replace(" ", "-").replace("/", "-")

    # Get date from section data
    date = section_data.get("date", "unknown") or "unknown"

    # Get amount info - could be sterling_total or calculated from mileage
    if active_section == "travel_mileage":
        miles = section_data.get("miles", 0) or 0
        cost_per_mile = section_data.get("cost_per_mile", 0) or 0
        amount = round(miles * cost_per_mile, 2)
        currency = "GBP"
    else:
        sterling_total = section_data.get("sterling_total", 0) or 0
        foreign_currency = section_data.get("foreign_currency", "")
        amount = sterling_total
        currency = "GBP"

    # Get description or location for vendor-like field
    if active_section == "travel_general":
        vendor = section_data.get("from_location", "") or section_data.get("to_location", "") or "travel"
    elif active_section == "travel_mileage":
        vendor = section_data.get("from_location", "") or section_data.get("to_location", "") or "mileage"
    elif active_section == "hospitality":
        vendor = section_data.get("principal_guest", "") or section_data.get("organisation", "") or "hospitality"
    else:
        vendor = section_data.get("description", "") or "expense"

    # Clean vendor name for filename
    vendor_clean = re.sub(r"[^\w\s-]", "", str(vendor))[:20].strip().replace(" ", "-").lower()
    if not vendor_clean:
        vendor_clean = "expense"

    # Get original extension
    ext = Path(original_filename).suffix.lower()
    if ext in [".heic", ".pdf"]:
        ext = ".png"  # We converted these

    return f"{expense_type}_{date}_{vendor_clean}_{amount}{currency}{ext}"


def fill_excel_template(
    template_path: str, header_info: dict, receipts: List[dict]
) -> bytes:
    """
    Fill the Excel template with header info and receipts.
    Returns the filled workbook as bytes.
    Uses new type-specific field structure.
    """
    wb = load_workbook(template_path)
    ws = wb["Portrait"]

    # Fill header info
    if header_info.get("name"):
        set_cell_value(ws, "C3", header_info["name"])
    if header_info.get("cid"):
        set_cell_value(ws, "C4", header_info["cid"])
    if header_info.get("dob"):
        set_cell_value(ws, "G5", header_info["dob"])
    if header_info.get("address"):
        set_cell_value(ws, "G6", header_info["address"])
    if header_info.get("postcode"):
        set_cell_value(ws, "I10", header_info["postcode"])
    if header_info.get("bank_name"):
        set_cell_value(ws, "K5", header_info["bank_name"])
    if header_info.get("bank_branch"):
        set_cell_value(ws, "K6", header_info["bank_branch"])
    if header_info.get("sort_code"):
        set_cell_value(ws, "K7", header_info["sort_code"])
    if header_info.get("account_number"):
        set_cell_value(ws, "K9", header_info["account_number"])

    # Get exchange rate from header info
    exchange_rate = float(header_info.get("exchange_rate", 1.0))

    # Track which rows we've used
    travel_idx = 0
    mileage_idx = 0
    hospitality_idx = 0
    other_idx = 0

    for receipt in receipts:
        parsed = receipt.get("parsed", {})
        active_section = parsed.get("active_section", "other")
        fields = parsed.get("fields", {})
        section_data = fields.get(active_section, {})
        excel_section = SECTION_TO_EXCEL.get(active_section, "other")

        if excel_section == "travel":
            if travel_idx < len(TRAVEL_ROWS):
                row = TRAVEL_ROWS[travel_idx]
                fill_travel_row(ws, row, section_data, exchange_rate)
                travel_idx += 1

        elif excel_section == "mileage":
            if mileage_idx < len(MILEAGE_ROWS):
                row = MILEAGE_ROWS[mileage_idx]
                fill_mileage_row(ws, row, section_data)
                mileage_idx += 1

        elif excel_section == "hospitality":
            if hospitality_idx < len(HOSPITALITY_ROWS):
                row = HOSPITALITY_ROWS[hospitality_idx]
                fill_hospitality_row(ws, row, section_data, exchange_rate)
                hospitality_idx += 1

        else:  # other
            if other_idx < len(OTHER_ROWS):
                row = OTHER_ROWS[other_idx]
                fill_other_row(ws, row, section_data, exchange_rate)
                other_idx += 1

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def fill_travel_row(ws, row: int, section_data: dict, exchange_rate: float = 1.0):
    """Fill a travel general section row using new field structure."""
    set_cell_value(ws, f"C{row}", section_data.get("date") or "")
    set_cell_value(ws, f"D{row}", section_data.get("mode") or "")  # Mode dropdown
    set_cell_value(ws, f"E{row}", "Yes" if section_data.get("is_return") else "")  # Return?
    set_cell_value(ws, f"F{row}", section_data.get("from_location") or "")  # From
    to_loc = section_data.get("to_location") or ""
    set_cell_value(ws, f"G{row}", to_loc[:30] if to_loc else "")  # To (truncated)

    # Currency handling
    foreign_currency = section_data.get("foreign_currency")  # e.g., "50.00 USD" or None
    sterling_total = section_data.get("sterling_total") or 0

    if foreign_currency:
        set_cell_value(ws, f"I{row}", foreign_currency)  # Foreign column
        set_cell_value(ws, f"J{row}", sterling_total)  # Sterling total
    else:
        set_cell_value(ws, f"J{row}", sterling_total)  # Sterling total only

    # Non UK/EU checkbox (column K)
    if section_data.get("is_non_uk_eu", False):
        set_cell_value(ws, f"K{row}", True)


def fill_hospitality_row(ws, row: int, section_data: dict, exchange_rate: float = 1.0):
    """Fill a hospitality section row using new field structure."""
    set_cell_value(ws, f"C{row}", section_data.get("date") or "")
    set_cell_value(ws, f"D{row}", section_data.get("principal_guest") or "")  # Name of principal guest
    org = section_data.get("organisation") or ""
    set_cell_value(ws, f"E{row}", org[:40] if org else "")  # Organisation (truncated)
    set_cell_value(ws, f"G{row}", section_data.get("total_numbers") or 1)  # Total numbers present

    # Currency handling
    foreign_currency = section_data.get("foreign_currency")
    sterling_total = section_data.get("sterling_total") or 0

    if foreign_currency:
        set_cell_value(ws, f"I{row}", foreign_currency)  # Foreign column
        set_cell_value(ws, f"J{row}", sterling_total)  # Sterling total
    else:
        set_cell_value(ws, f"J{row}", sterling_total)

    # Non-college staff present checkbox
    if section_data.get("non_college_staff", False):
        set_cell_value(ws, f"K{row}", True)


def fill_other_row(ws, row: int, section_data: dict, exchange_rate: float = 1.0):
    """Fill a subsistence/other section row using new field structure."""
    set_cell_value(ws, f"C{row}", section_data.get("date") or "")
    set_cell_value(ws, f"D{row}", section_data.get("expense_type") or "")  # Expense type dropdown
    desc = section_data.get("description") or ""
    set_cell_value(ws, f"E{row}", desc[:50] if desc else "")  # Description (truncated)

    # Currency handling
    foreign_currency = section_data.get("foreign_currency")
    sterling_total = section_data.get("sterling_total") or 0

    if foreign_currency:
        set_cell_value(ws, f"I{row}", foreign_currency)  # Foreign column
        set_cell_value(ws, f"J{row}", sterling_total)  # Sterling total
    else:
        set_cell_value(ws, f"J{row}", sterling_total)

    # Non UK/EU checkbox (column K)
    if section_data.get("is_non_uk_eu", False):
        set_cell_value(ws, f"K{row}", True)


def fill_mileage_row(ws, row: int, section_data: dict):
    """Fill a car mileage section row using new field structure."""
    set_cell_value(ws, f"C{row}", section_data.get("date") or "")
    set_cell_value(ws, f"D{row}", section_data.get("miles") or "")  # Number of miles
    set_cell_value(ws, f"E{row}", "Yes" if section_data.get("is_return") else "")  # Return?
    set_cell_value(ws, f"F{row}", section_data.get("from_location") or "")  # From
    set_cell_value(ws, f"G{row}", section_data.get("to_location") or "")  # To
    set_cell_value(ws, f"H{row}", section_data.get("cost_per_mile") or "")  # Cost per mile

    # Calculate total (miles * cost_per_mile) for column J
    miles = section_data.get("miles") or 0
    cost_per_mile = section_data.get("cost_per_mile") or 0
    if miles and cost_per_mile:
        total = round(miles * cost_per_mile, 2)
        set_cell_value(ws, f"J{row}", total)


def split_into_batches(receipts: List[dict]) -> List[List[dict]]:
    """
    Split receipts into batches that fit within Excel template row limits.
    Each batch respects: MAX_TRAVEL=6, MAX_MILEAGE=4, MAX_HOSPITALITY=4, MAX_OTHER=7
    Uses new active_section field structure.
    """
    # Group receipts by section
    travel = []
    mileage = []
    hospitality = []
    other = []

    for receipt in receipts:
        parsed = receipt.get("parsed", {})
        active_section = parsed.get("active_section", "other")
        excel_section = SECTION_TO_EXCEL.get(active_section, "other")

        if excel_section == "travel":
            travel.append(receipt)
        elif excel_section == "mileage":
            mileage.append(receipt)
        elif excel_section == "hospitality":
            hospitality.append(receipt)
        else:
            other.append(receipt)

    # Calculate how many batches needed
    num_batches = max(
        1,
        (len(travel) + MAX_TRAVEL - 1) // MAX_TRAVEL if travel else 1,
        (len(mileage) + MAX_MILEAGE - 1) // MAX_MILEAGE if mileage else 1,
        (len(hospitality) + MAX_HOSPITALITY - 1) // MAX_HOSPITALITY if hospitality else 1,
        (len(other) + MAX_OTHER - 1) // MAX_OTHER if other else 1,
    )

    # Distribute receipts across batches
    batches = [[] for _ in range(num_batches)]

    for i, receipt in enumerate(travel):
        batch_idx = i // MAX_TRAVEL
        if batch_idx < num_batches:
            batches[batch_idx].append(receipt)

    for i, receipt in enumerate(mileage):
        batch_idx = i // MAX_MILEAGE
        if batch_idx < num_batches:
            batches[batch_idx].append(receipt)

    for i, receipt in enumerate(hospitality):
        batch_idx = i // MAX_HOSPITALITY
        if batch_idx < num_batches:
            batches[batch_idx].append(receipt)

    for i, receipt in enumerate(other):
        batch_idx = i // MAX_OTHER
        if batch_idx < num_batches:
            batches[batch_idx].append(receipt)

    # Filter out empty batches
    return [b for b in batches if b]


def create_output_zip(
    template_path: str,
    header_info: dict,
    receipts: List[dict],
    surname: str = "Expense"
) -> bytes:
    """
    Create a ZIP file containing:
    - Filled Excel form(s)
    - Renamed receipt images
    If receipts exceed template row limits, creates multiple folders (expense-1, expense-2, etc.)
    """
    buffer = io.BytesIO()

    # Split receipts into batches that fit the template
    batches = split_into_batches(receipts)

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for batch_num, batch in enumerate(batches, 1):
            # Determine folder prefix
            if len(batches) > 1:
                folder_prefix = f"expense-{batch_num}/"
            else:
                folder_prefix = ""

            # Generate filled Excel for this batch
            excel_bytes = fill_excel_template(template_path, header_info, batch)
            excel_filename = f"{folder_prefix}{surname}_E1-expense-form.xlsx"
            zf.writestr(excel_filename, excel_bytes)

            # Add renamed receipts for this batch
            for receipt in batch:
                original_name = receipt.get("filename", "receipt.png")
                parsed = receipt.get("parsed", {})
                new_name = generate_expense_filename(parsed, original_name)

                # Decode and add image
                image_b64 = receipt.get("image_base64", "")
                if image_b64:
                    image_bytes = base64.b64decode(image_b64)
                    zf.writestr(f"{folder_prefix}receipts/{new_name}", image_bytes)

    buffer.seek(0)
    return buffer.read()
